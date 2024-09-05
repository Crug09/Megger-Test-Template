import os
import datetime
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.lang import Builder

# Define different screens
class MainScreen(Screen):
    def process_input(self):

        customer_name = self.ids.customer_name.text
        customer_name = customer_name.replace(' ', '_')
        customer_name = customer_name.title()
        if customer_name:
            phase_labels = ['Phase 1', 'Phase 2', 'Phase 3', ' ']
            underground_wire = self.ids.underground_wire.active
            tower_disconnect_switch =self.ids.tower_disconnect_switch.active
            tower_wire = self.ids.tower_wire.active
            junction_box = self.ids.junction_box.active
            slip_rings = self.ids.slip_rings.active
            brush_block = self.ids.brush_block.active
            altenator = self.ids.altenator.active

            phase_1 = self.ids.phase_1.text
            phase_2 = self.ids.phase_2.text
            phase_3 = self.ids.phase_3.text

            if phase_1 and phase_2 and phase_3:


                megger_test = []
                if underground_wire:
                    megger_test.append('Underground_Wire')
                if tower_disconnect_switch:
                    megger_test.append('Tower_Disconnect_Switch')
                if tower_wire:
                    megger_test.append('Tower_Wire')
                if junction_box:
                    megger_test.append('Junction_Box')
                if slip_rings:
                    megger_test.append('Slip_Rings')
                if brush_block:
                    megger_test.append('Brush_Block')
                if altenator:
                    megger_test.append('Altenator')

                megger_test = '__'.join(megger_test)

                if megger_test != '':

                    today_date = datetime.datetime.now().strftime('%Y-%m-%d')
                    file_name = f"{customer_name}_Megger_Test_{today_date}.xlsx"

                    phase_results = [int(phase_1), int(phase_2), int(phase_3), ' ']
                    
                    if os.path.exists(file_name):
                        workbook = load_workbook(file_name)
                        sheet = workbook.active
                        workbook.save(file_name)
                    else:
                        workbook = Workbook()
                        sheet = workbook.active
                        sheet.title = "Sheet1"
                        workbook.save(file_name)

                    workbook = load_workbook(file_name)

                    sheet = workbook.active
                    row_number = 1
                    row = sheet[row_number]
                    components = [cell.value for cell in row if cell.value is not None]

                    if megger_test in components:
                        pass                        
                    else:
                        last_column = sheet.max_column
                        new_column = last_column + 1
                        cell = sheet.cell(row=row_number, column=new_column)
                        cell.value = megger_test
                        workbook.save(file_name) 
        
                    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red background
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green background

                    target_column = None
                    for cell in sheet[1]: 
                        if cell.value == megger_test:
                            target_column = cell.column
                            break
                    if target_column is None:
                        self.ids.megger_test_id.text = 'Error has occured'
                    else:
                        start_row = 1
                        while sheet.cell(row=start_row, column=target_column).value is not None:
                            start_row += 1

                        for i, (item, phase) in enumerate(zip(phase_results, phase_labels)):
                            phase_cell = sheet.cell(row=start_row + i, column=1)
                            phase_cell.value = phase
                            
                            data_cell = sheet.cell(row=start_row + i, column=target_column)
                            data_cell.value = item

                        for row in range(start_row, start_row + len(phase_results)): 
                            cell = sheet.cell(row=row, column=target_column)
                            
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                if 0 <= cell.value <= 349:
                                    cell.fill = red_fill
                                elif 350 <= cell.value <= 524:
                                    cell.fill = yellow_fill
                                elif 525 <= cell.value <= 550:
                                    cell.fill = green_fill

                        workbook.save(file_name)
                        megger_test = megger_test.replace('__', ' and ')
                        megger_test = megger_test.replace('_', ' ')
                        self.ids.megger_test_id.text = f'Added:\n{megger_test}\nWith the data being:{phase_results[:3]}'
                else:
                    self.ids.megger_test_id.text = 'Enter a component'

            else:
                self.ids.megger_test_id.text = 'Enter all phases'
        else:
            self.ids.megger_test_id.text = 'Enter name of customer'



        
    def clear_inputs(self):
        self.ids.customer_name.text = ''
        self.ids.underground_wire.active = False
        self.ids.tower_disconnect_switch.active = False
        self.ids.tower_wire.active = False
        self.ids.junction_box.active = False
        self.ids.slip_rings.active = False
        self.ids.brush_block.active = False
        self.ids.altenator.active = False
        self.ids.phase_1.text = ''
        self.ids.phase_2.text = ''
        self.ids.phase_3.text = ''




class NamesScreen(Screen):
    def update_file_list(self, file_names):
        if file_names:
            self.ids.file_list_label.text = '\n'.join(file_names)
        else:
            self.ids.file_list_label.text = "No previous test files found."

class ScreenManagement(ScreenManager):
    pass

class MyKivyApp(App):
    def build(self):
        Builder.load_file('main.kv')
        # Builder.load_file('name.kv')
        
        sm = ScreenManagement()
        sm.add_widget(MainScreen(name='main_screen'))
        sm.add_widget(NamesScreen(name='name_screen'))
        return sm

    
    def show_excel_files(self):
        directory_path = '.'  
        file_names = [f for f in os.listdir(directory_path) if f.endswith('.xlsx')]
        customer_names = []
        x = 0
        for i in file_names:
            base_name = os.path.splitext(file_names[x])[0]
            parts = re.split('[._]', base_name)
            customer_names.append(parts[0] + " " + parts[1] + " " + parts[4] )
            x+=1
        name_screen = self.root.get_screen('name_screen')
        name_screen.update_file_list(customer_names)
        self.root.current = 'name_screen'
        

if __name__ == '__main__':
    MyKivyApp().run()
